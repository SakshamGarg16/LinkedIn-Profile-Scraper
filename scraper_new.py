import csv
import time
import random
import requests
from googlesearch import search
from openpyxl import Workbook
from openpyxl.styles import Font

# Define job roles to search
search_terms = ["CTO", "Co-Founder", "Founder","Employee"]

# Function to fetch LinkedIn profiles using Google Search
def get_linkedin_profiles(company_name):
    print(f" Searching LinkedIn profiles for {company_name}...")
    profiles = []
    seen_profiles = set()

    for role in search_terms:
        query = f"{company_name} {role} site:linkedin.com/in"
        try:
            results = search(query, num_results=2)  # No. of results needed

            for link in results:
                if "linkedin.com/in/" in link:  
                    profile_name = link.split("/in/")[-1].split("?")[0].replace("-", " ").title()
                    if profile_name and profile_name not in seen_profiles:
                        seen_profiles.add(profile_name)
                        profiles.append((profile_name, company_name, role, link))
                        print(f"Found: {profile_name} - {role} - {link}")

            time.sleep(random.uniform(3, 6))

        except Exception as e:
            print(f"Error fetching results for {query}: {e}")

    return profiles

input_file = "input_companies.csv"
output_file = "linkedin_profiles.xlsx"

with open(input_file, "r") as csvfile:
    companies = [row[0] for row in csv.reader(csvfile)]
    total_companies = len(companies)

wb = Workbook()
ws = wb.active
ws.title = "LinkedIn Profiles"
ws.append(["Profile Name", "Company Name", "Role", "LinkedIn URL"])

for i, company in enumerate(companies):
    profiles = get_linkedin_profiles(company)

    for profile in profiles:
        ws.append([profile[0], profile[1], profile[2], profile[3]])
        cell = ws.cell(row=ws.max_row, column=4)
        cell.font = Font(color="0000FF", underline="single") 

    progress = int((i + 1) / total_companies * 40)
    print(f"[{'#' * progress}{' ' * (40 - progress)}] {i+1}/{total_companies}", end="\r")

wb.save(output_file)
print(f"\n LinkedIn profiles saved to {output_file}")
